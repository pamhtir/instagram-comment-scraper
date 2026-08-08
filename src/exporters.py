"""Lossless raw snapshots and client-ready processed exports."""

import json
from dataclasses import asdict, is_dataclass
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


NAVY, BLUE, PALE, WHITE, GREY = "172B4D", "1F6FEB", "EAF2FF", "FFFFFF", "667085"


def export_raw_records(records: Iterable[Any], output_path: Path) -> Path:
    """Write the scraper output unchanged as UTF-8 JSON Lines before cleaning."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="\n") as handle:
        for record in records:
            if hasattr(record, "to_dict"):
                payload = record.to_dict()
            elif is_dataclass(record):
                payload = asdict(record)
            elif isinstance(record, dict):
                payload = record
            else:
                raise TypeError(f"Unsupported raw record type: {type(record).__name__}")
            handle.write(json.dumps(payload, ensure_ascii=False, default=str) + "\n")
    return output_path


def export_results(df: pd.DataFrame, metrics: dict[str, int], output_stem: Path) -> tuple[Path, Path]:
    output_stem.parent.mkdir(parents=True, exist_ok=True)
    csv_path, xlsx_path = output_stem.with_suffix(".csv"), output_stem.with_suffix(".xlsx")
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    unique_users = int(df["username"].nunique()) if "username" in df else 0
    replies = int(df["is_reply"].fillna(False).astype(bool).sum()) if "is_reply" in df else 0
    summary = pd.DataFrame({
        "Metric": ["Raw records", "Invalid records", "Duplicates removed", "Final comments", "Unique users", "Replies"],
        "Value": [metrics["raw"], metrics["invalid"], metrics["duplicates"], metrics["final"], unique_users, replies],
    })
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False, startrow=3)
        df.to_excel(writer, sheet_name="Comments", index=False)
        summary_ws = writer.book["Summary"]
        summary_ws["A1"] = "INSTAGRAM COMMENT EXPORT"
        summary_ws["A1"].font = Font(size=18, bold=True, color=WHITE)
        summary_ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
        summary_ws.merge_cells("A1:B2")
        summary_ws["A1"].alignment = Alignment(vertical="center")
        summary_ws.row_dimensions[1].height = 28
        for ws in (summary_ws, writer.book["Comments"]):
            ws.sheet_view.showGridLines = False
            ws.freeze_panes = "A5" if ws.title == "Summary" else "A2"
            header_row = 4 if ws.title == "Summary" else 1
            for cell in ws[header_row]:
                cell.font = Font(bold=True, color=WHITE)
                cell.fill = PatternFill("solid", fgColor=BLUE)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for row in ws.iter_rows(min_row=header_row + 1):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = Border(bottom=Side(style="hair", color="D0D5DD"))
            for cells in ws.iter_cols():
                letter = get_column_letter(cells[0].column)
                width = max(len(str(cell.value or "")) for cell in cells) + 2
                ws.column_dimensions[letter].width = min(max(width, 12), 55)
        if len(df):
            comments_ws = writer.book["Comments"]
            table = Table(displayName="CommentsTable", ref=comments_ws.dimensions)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
            comments_ws.add_table(table)
            comments_ws.column_dimensions["F"].width = 55
        summary_ws.column_dimensions["A"].width = 24
        summary_ws.column_dimensions["B"].width = 16
    return csv_path, xlsx_path
