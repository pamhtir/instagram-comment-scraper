import pandas as pd
from openpyxl import load_workbook

from src.cleaners import ORDERED_COLUMNS
import json

from src.exporters import export_raw_records, export_results
from src.models import CommentRecord


def test_exports_lossless_raw_jsonl(tmp_path):
    record = CommentRecord(
        platform="instagram",
        content_type="reel",
        source_url="https://www.instagram.com/reel/X/",
        comment_id="id1",
        username="alice",
        comment="Unicode survives: Việt Nam 😮",
        scraped_at_utc="2026-01-01T00:00:00+00:00",
    )
    path = export_raw_records([record], tmp_path / "raw" / "records.jsonl")
    assert path.exists()
    assert json.loads(path.read_text(encoding="utf-8")) == record.to_dict()


def test_exports_empty_raw_snapshot(tmp_path):
    path = export_raw_records([], tmp_path / "raw" / "empty.jsonl")
    assert path.exists()
    assert path.read_text(encoding="utf-8") == ""


def test_exports_csv_and_workbook(tmp_path):
    row = ["instagram", "reel", "https://www.instagram.com/reel/X/", "id1", "alice", "Useful!", False, "", "2026-01-01T00:00:00+00:00"]
    df = pd.DataFrame([row], columns=ORDERED_COLUMNS)
    csv_path, xlsx_path = export_results(df, {"raw": 1, "invalid": 0, "duplicates": 0, "final": 1}, tmp_path / "result")
    assert csv_path.exists() and xlsx_path.exists()
    workbook = load_workbook(xlsx_path, read_only=False)
    assert workbook.sheetnames == ["Summary", "Comments"]
    assert workbook["Comments"].max_row == 2


def test_exports_empty_result(tmp_path):
    df = pd.DataFrame(columns=ORDERED_COLUMNS)
    _, xlsx_path = export_results(df, {"raw": 0, "invalid": 0, "duplicates": 0, "final": 0}, tmp_path / "empty")
    assert xlsx_path.exists()
